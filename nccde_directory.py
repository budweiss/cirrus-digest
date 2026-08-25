#!/usr/bin/env python3
"""
nccde_association_directory.py — pull New Castle County's PUBLIC community
association directory (maintenance corporations + civic associations).

S77 (2026-08-25). This is the source we told Bill in July we would find and
then didn't: ~224 New Castle County associations WITH a named officer or
management company, and for most of them an email and/or phone. It is free,
needs no account, and is served as plain server-rendered HTML — no portal
login, no deed search, no per-page fee.

Why it matters more than the parcel data we had: `nccde_subdivisions_clean.json`
gives community NAMES and acreage. This gives a PERSON TO CONTACT. It also
separates self-managed associations (Bill's prospects) from ones already held
by a management company (his competitors) — Knight Property Services itself
appears here, on Enclave at Odessa.

Emails are obfuscated on the page by a tiny inline script that concatenates
user + domain at render time, so a plain HTML scrape misses them; we
reconstruct from the script's own variables.

Two modes:
  -o FILE            dump the parsed directory as JSON (the lead workbook)
  --enrich PROJECT   match it against an entity_kb project and write the
                     officer/management details onto the entities we already
                     track. DRY RUN by default; --apply writes.

Matching is deliberately strict: normalised EXACT name only. A near match is
reported, never applied. Attaching the wrong president to the wrong community
would put a stranger's name in front of a client, which is worse than a gap.
"""

import argparse
import html
import json
import re
import sys
import urllib.request

URL = ("https://www.newcastlede.gov/BusinessDirectoryII.aspx"
       "?lngBusinessCategoryID=31,32&ysnShowAll=1")
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"

# Terminate each item at its own closing </div>, NOT at the next item's name:
# the last row on the page has no following name, and a lookahead for one
# silently dropped it (caught by the count guard in main(), 2026-08-25).
ROW_RX = re.compile(
    r'<span style="font-weight: bold;">(?P<name>[^<]+)</span>(?P<body>.*?)</div>',
    re.S)
EMAIL_RX = re.compile(r"var w = '([^']*)'\s*var x = '([^']*)'", re.S)
PHONE_RX = re.compile(r'Phone:\s*<a href="tel:[^"]*">([^<]+)</a>')
LINK_RX = re.compile(r'Link:\s*<a href="([^"]+)"')


# The directory's own "<Company>, MC" suffix is the explicit marker for a
# professionally managed community -- but it is NOT applied consistently.
# "Neighborhood Resources LLC, Steve Blanchies" carries no MC and covers ~10
# communities; on the suffix alone every one of them landed in Bill's PROSPECT
# list, which would have had him cold-pitching a competitor at that
# competitor's own email address (caught 2026-08-25 by reading the workbook
# before it went out, not by any check).
_COMPANY_RX = re.compile(
    r"\b(LLC|L\.L\.C|INC|CORP|CO|COMPANY|MANAGEMENT|MGMT|PROPERTIES|PROPERTY|"
    r"SERVICES|REALTY|GROUP|ASSOCIA|PARTNERS|ADVISORS)\b", re.IGNORECASE)


def _classify_contact(contact: str):
    """(is_managed, why). The 'why' is carried into the workbook so a judgement
    call is auditable by the person acting on it, rather than a bare boolean."""
    c = (contact or "").strip()
    if re.search(r",\s*MC$", c):
        return True, "directory 'MC' marker"
    if _COMPANY_RX.search(c):
        return True, "contact is a company name"
    if re.search(r",\s*(President|Treasurer|Chair|Vice President|Secretary)\b",
                 c, re.IGNORECASE):
        return False, "named officer listed"
    return False, "no company or officer marker — UNVERIFIED, check before use"


def fetch(url: str = URL) -> str:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=60) as r:
        return r.read().decode("utf-8", errors="replace")


def parse(page: str) -> list:
    out = []
    for m in ROW_RX.finditer(page):
        body = m.group("body")
        # Strip the inline scripts before reading the address lines, else the
        # email-obfuscation JS lands in the address.
        clean = re.sub(r"<script.*?</script>", "", body, flags=re.S)
        parts = [html.unescape(p).replace("\xa0", " ").strip()
                 for p in re.split(r"<br\s*/?>", clean)]
        parts = [p for p in (re.sub(r"<[^>]+>", "", p).strip() for p in parts) if p]
        parts = [p for p in parts if not p.startswith(("Phone:", "Email:", "Link:"))]

        em = EMAIL_RX.search(body)
        ph = PHONE_RX.search(body)
        ln = LINK_RX.search(body)
        contact = parts[0] if parts else ""
        managed, basis = _classify_contact(contact)
        rec = {
            "name": html.unescape(m.group("name")).strip(),
            "contact": contact,
            "managed": managed,
            "managed_basis": basis,
            "address": ", ".join(parts[1:]),
            "phone": ph.group(1).strip() if ph else "",
            "email": f"{em.group(1)}@{em.group(2)}" if em else "",
            "website": ln.group(1) if ln else "",
        }
        out.append(rec)
    return out


SOURCE_URL = URL
CAVEAT = (
    "New Castle County publishes this directory with an explicit disclaimer "
    "that it makes no warranty as to accuracy, and entries go stale when a "
    "board turns over. Treat every contact as a strong starting point, not a "
    "verified contact list — confirm before anything goes on a contract.")


def workbook(recs: list, path: str) -> str:
    """Bill's lead workbook. Ordered the way it will actually be worked:
    self-managed communities WITH a reachable contact first, because those are
    the ones he can act on today. Competitors get their own tab -- knowing
    Aspen holds 11 is as useful as any single prospect."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    def sheet(wb, title, headers, rows):
        ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
        ws.title = title
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append(r)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i, h in enumerate(headers, 1):
            width = max([len(str(h))] + [len(str(r[i - 1] or "")) for r in rows] or [10])
            ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 12), 60)
        return ws

    prospects = [r for r in recs if not r["managed"]]
    # reachable first: email beats phone beats neither
    prospects.sort(key=lambda r: (not r["email"], not r["phone"], r["name"].lower()))
    managed = sorted((r for r in recs if r["managed"]),
                     key=lambda r: (r["contact"].lower(), r["name"].lower()))

    wb = Workbook()
    sheet(wb, "Prospects (self-managed)",
          ["Association", "Contact", "Email", "Phone", "Address", "Website",
           "Why we think it is self-managed"],
          [[r["name"], r["contact"], r["email"], r["phone"], r["address"],
            r["website"], r.get("managed_basis", "")] for r in prospects])
    sheet(wb, "Already managed",
          ["Association", "Management company", "Email", "Phone", "Address",
           "How identified"],
          [[r["name"], r["contact"], r["email"], r["phone"], r["address"],
            r.get("managed_basis", "")] for r in managed])

    from collections import Counter
    counts = Counter(r["contact"] for r in managed)
    sheet(wb, "Who holds what",
          ["Management company", "Communities held"],
          [[k, v] for k, v in counts.most_common()])

    ws = wb.create_sheet("Source & caveats")
    for row in [
        ["Source", "New Castle County Community Association Directory"],
        ["URL", SOURCE_URL],
        ["Access", "Public. No account required."],
        ["Total associations", len(recs)],
        ["Self-managed (prospects)", len(prospects)],
        ["Professionally managed", len(managed)],
        ["With an email on file", sum(1 for r in recs if r["email"])],
        ["With a phone on file", sum(1 for r in recs if r["phone"])],
        ["Caveat", CAVEAT],
    ]:
        ws.append(row)
    ws["A1"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 100

    import os
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)
    return path


def build_workbook(path: str) -> str:
    """fetch -> parse -> workbook, with the same count guard as main()."""
    page = fetch()
    recs = parse(page)
    expected = re.search(r"of (\d+) Listing", page)
    if not recs or (expected and len(recs) != int(expected.group(1))):
        raise RuntimeError(
            f"directory parse mismatch: page says "
            f"{expected.group(1) if expected else '?'}, parsed {len(recs)}")
    return workbook(recs, path)


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()


DIRECTORY_SOURCE = "nccde_association_directory"
NCC = "New Castle"


def _county_ok(entity: dict) -> bool:
    """Can a NEW CASTLE directory row bind to this entity at all?

    The one rule that matters in this file, and it now lives in ONE place.
    S78 added it to enrich()'s update path after New Castle's "Hunters Ridge"
    overwrote the researched president of KENT's "Hunters Ridge" -- a warm,
    tier-A lead. It was NOT added to the two name-only sets below, and that
    half-fix cost the client a row: the same directory row was refused an
    update (right) and then counted as "already present" by import_new (wrong),
    while enrich's closing "rows with no entity here" line -- the one that
    exists to say what was NOT covered -- reported 0.

    A row may be rejected as the wrong county, or counted as already covered.
    It must never be both. An entity with no county on file is still eligible;
    only a KNOWN other county disqualifies.
    """
    c = str((entity.get("state") or {}).get("county", "")).strip()
    return not c or _norm(c) == _norm(NCC)


def _bindable_names(entities: list) -> set:
    """Normalised names a New Castle directory row may legitimately match."""
    return {_norm(e["name"]) for e in entities if _county_ok(e)}


def _free_slug(name: str, taken: set) -> str:
    """A slug that is not already somebody else's.

    slugify() is derived from the name alone, so two same-named associations in
    two counties yield the SAME slug -- and upsert_entity() on a taken slug
    UPDATES that entity rather than creating one. Importing New Castle's
    "Hunters Ridge" under the bare slug would therefore have overwritten Kent's,
    which is precisely the corruption the county guard was added to stop. The
    guard would have been bypassed by the path meant to honour it.
    """
    import entity_kb
    slug = entity_kb.slugify(name)
    if slug not in taken:
        return slug
    cand, n = slug + "-new-castle", 2
    while cand in taken:
        cand, n = f"{slug}-new-castle-{n}", n + 1
    return cand


def import_new(project: str, recs: list, apply: bool) -> dict:
    """Create entities for directory rows this project does not have yet.

    Kept SEPARATE from enrich() and separately tagged (Buddy, 2026-08-25): the
    project's existing leads come from parcel/plat provenance, and a lead that
    silently appears from a second source with no marker would make
    `confidence_basis` meaningless and would surface in a client digest with no
    explanation of where it came from.
    """
    import entity_kb

    entities = entity_kb.list_entities(project)
    # See _county_ok(): a same-named entity in ANOTHER county is not this
    # association, so it must not make this row look already-covered.
    have = _bindable_names(entities)
    taken = {e["slug"] for e in entities}
    created, skipped, renamed = [], 0, []
    for r in recs:
        if _norm(r["name"]) in have:
            skipped += 1
            continue
        slug = _free_slug(r["name"], taken)
        if slug != entity_kb.slugify(r["name"]):
            renamed.append(f"{r['name']} -> {slug}")
        taken.add(slug)
        fields = {
            "source": DIRECTORY_SOURCE,
            "county": "New Castle",
            "type": "HOA",
            "directory_source": ("New Castle County Community Association "
                                 "Directory (public, county-maintained)"),
            "confidence_basis": (
                "Listed in the New Castle County Community Association "
                "Directory. The association and its contact are county-published; "
                "NO parcel, acreage or owner data has been matched to it yet."),
            "lead_confidence": "medium",
        }
        if r["managed"]:
            fields["current_mgmt_co"] = r["contact"]
            fields["mgmt_status"] = (f"professionally managed ({r['contact']}) "
                                     f"[{r.get('managed_basis', '')}]")
        else:
            fields["mgmt_status"] = ("self-managed (no management company listed) "
                                     f"[{r.get('managed_basis', '')}]")
            if r["contact"]:
                fields["board_contact"] = r["contact"]
        for k, v in (("board_email", r["email"]), ("board_phone", r["phone"]),
                     ("website", r["website"]), ("mailing_address", r["address"])):
            if v:
                fields[k] = v

        if not apply:
            created.append(r["name"])
            continue
        entity_kb.upsert_entity(project, slug, r["name"],
                                entity_type="hoa", fields=fields,
                                lead_state="new")
        created.append(r["name"])

    print(f"{'IMPORTED' if apply else 'WOULD IMPORT'}: {len(created)} new "
          f"(already present: {skipped})")
    for n in created[:15]:
        print(f"    + {n}")
    if len(created) > 15:
        print(f"    ... and {len(created) - 15} more")
    # Never silent: a slug that had to be disambiguated means this project now
    # holds two same-named associations in different counties. Say so.
    if renamed:
        print(f"  name already used by another county's entity, "
              f"imported under a distinct slug: {len(renamed)}")
        for x in renamed:
            print(f"    ~ {x}")
    return {"created": len(created), "skipped": skipped, "renamed": len(renamed)}


def enrich(project: str, recs: list, apply: bool) -> int:
    """Write directory officer/management details onto entities we already
    track. Only ever UPDATES an existing entity -- the directory is a contact
    source, not a source of new leads, and inventing entities from it would mix
    two provenances in one project."""
    import entity_kb

    by_norm = {}
    for r in recs:
        by_norm.setdefault(_norm(r["name"]), []).append(r)

    entities = entity_kb.list_entities(project)
    matched, changed, ambiguous, unmatched = 0, 0, [], 0
    wrong_county = []
    for e in entities:
        hits = by_norm.get(_norm(e["name"]), [])
        if not hits:
            unmatched += 1
            continue
        # 2026-08-25 (S78) — THIS IS A NEW CASTLE COUNTY DIRECTORY. An exact NAME
        # match is not an identity match: Delaware reuses community names across
        # counties, and this project already rejects same-name-different-STATE
        # hits for exactly that reason. Without a county guard the directory row
        # for New Castle's "Hunters Ridge" overwrote the researched president of
        # KENT's "Hunters Ridge" -- a warm, tier-A lead -- with a stranger's
        # name. Mailing that board would have reached the wrong association.
        # An entity with no county on file is still eligible; only a KNOWN other
        # county is disqualifying.
        ent_county = str((e.get("state") or {}).get("county", "")).strip()
        if ent_county and _norm(ent_county) != _norm("New Castle"):
            wrong_county.append(f"{e['name']} (on file as {ent_county})")
            continue
        if len(hits) > 1:
            ambiguous.append(e["name"])
            continue
        r = hits[0]
        matched += 1

        fields = {}
        if r["managed"]:
            fields["current_mgmt_co"] = r["contact"]
            fields["mgmt_status"] = (f"professionally managed ({r['contact']}) "
                                     f"[{r.get('managed_basis', '')}]")
        else:
            fields["mgmt_status"] = ("self-managed (no management company listed) "
                                     f"[{r.get('managed_basis', '')}]")
            if r["contact"]:
                fields["board_contact"] = r["contact"]
        # Never blank out something we already hold with something we don't.
        if r["email"]:
            fields["board_email"] = r["email"]
        if r["phone"]:
            fields["board_phone"] = r["phone"]
        if r["website"]:
            fields["website"] = r["website"]
        fields["directory_source"] = ("New Castle County Community Association "
                                      "Directory (public, county-maintained)")

        if not apply:
            # DIFF against what is stored, don't just count the match. A dry run
            # that reports "109 would change" every time, including straight
            # after a successful apply, can never tell you there is nothing to
            # do -- and a report that is never quiet gets ignored (T9).
            cur = (entity_kb.get_entity(project, e["slug"]) or {}).get("state", {})
            diff = sorted(k for k, v in fields.items() if cur.get(k) != v)
            if diff:
                changed += 1
                print(f"  WOULD UPDATE {e['name']}: {diff}")
            continue
        res = entity_kb.upsert_entity(project, e["slug"], e["name"], fields=fields)
        if res.get("changed_fields"):
            changed += 1
            print(f"  {e['name']}: {res['changed_fields']}")

    print(f"\n{'APPLIED' if apply else 'DRY RUN'} — project '{project}'")
    print(f"  entities in project : {len(entities)}")
    print(f"  matched to directory: {matched}")
    print(f"  with field changes  : {changed}"
          + ("  (already up to date)" if matched and not changed else ""))
    print(f"  ambiguous (skipped) : {len(ambiguous)}"
          + (f" -> {ambiguous[:5]}" if ambiguous else ""))
    print(f"  no directory entry  : {unmatched}")
    # Never silent: a skip the operator cannot see is how the Hunters Ridge
    # overwrite survived a dry run that reported "0 ambiguous" (T8/T9).
    print(f"  name matched but WRONG COUNTY (skipped): {len(wrong_county)}")
    for w in wrong_county:
        print(f"    - {w}")
    # Say what was NOT covered. A directory row we could not place is a lead
    # this project simply does not know about yet.
    # Same rule as import_new: an entity in another county does not "place" a
    # New Castle row. Using a bare name set here reported 0 uncovered rows while
    # Hunters Ridge sat uncovered -- a false all-clear on the one line whose
    # whole job is to say what was missed.
    placed = _bindable_names(entities)
    orphan = [r["name"] for r in recs if _norm(r["name"]) not in placed]
    print(f"  directory rows with no entity here: {len(orphan)}")
    for o in orphan[:10]:
        print(f"    ! {o}")
    if len(orphan) > 10:
        print(f"    ... and {len(orphan) - 10} more")
    return 0


def selftest() -> int:
    """Offline unit tests for the county/slug rules.

    S78. These three functions decide whether a client gains a lead, loses one,
    or has a researched contact overwritten by a stranger in another county, and
    until now nothing tested them -- the half-fix that dropped Hunters Ridge
    shipped through a dry run that printed a clean "0". Detection first: an
    instruction to be careful can be skipped, a check cannot.

    Needs entity_kb only for slugify(), and is LOUD if it cannot import it --
    a slug test that silently falls back to its own slugifier would be testing
    the wrong function and passing.
    """
    try:
        import entity_kb
    except Exception as e:
        print(f"  FAIL  cannot import entity_kb ({e}) -- run this on the box, "
              f"where the real slugify() lives")
        return 1

    bad = 0

    def check(label, ok):
        nonlocal bad
        print(f"  {'PASS' if ok else 'FAIL'}  {label}")
        if not ok:
            bad += 1

    ncc = {"name": "Back Creek", "slug": "back-creek",
           "state": {"county": "New Castle"}}
    kent = {"name": "Hunters Ridge", "slug": "hunters-ridge",
            "state": {"county": "Kent"}}
    blank = {"name": "Somewhere", "slug": "somewhere", "state": {}}

    check("a New Castle entity is bindable", _county_ok(ncc))
    check("a KENT entity is NOT bindable", not _county_ok(kent))
    check("an entity with no county on file is still bindable", _county_ok(blank))
    check("an entity with no state dict at all does not crash",
          _county_ok({"name": "x", "slug": "x"}))

    # The exact row that was lost. Kent's Hunters Ridge must not make the New
    # Castle directory row look already-covered.
    names = _bindable_names([ncc, kent, blank])
    check("a wrong-county name is excluded from the bindable set",
          _norm("Hunters Ridge") not in names)
    check("a right-county name is included", _norm("Back Creek") in names)
    check("an unknown-county name is included", _norm("Somewhere") in names)

    # ... and must not have its slug handed to the importer either.
    taken = {"hunters-ridge", "back-creek"}
    check("a free slug is used as-is", _free_slug("Brand New Place", taken)
          == entity_kb.slugify("Brand New Place"))
    hr = _free_slug("Hunters Ridge", taken)
    check("a taken slug is NOT reused (would overwrite the other county)",
          hr != "hunters-ridge")
    check("the disambiguated slug names the county", hr == "hunters-ridge-new-castle")
    check("a second collision keeps counting",
          _free_slug("Hunters Ridge", taken | {"hunters-ridge-new-castle"})
          == "hunters-ridge-new-castle-2")

    print("\nALL PASS" if not bad else f"\n{bad} FAILED")
    return 1 if bad else 0


def main() -> int:
    if "--selftest" in sys.argv:
        return selftest()
    ap = argparse.ArgumentParser()
    ap.add_argument("-o", "--out")
    ap.add_argument("--workbook", metavar="XLSX",
                    help="write the client lead workbook (.xlsx)")
    ap.add_argument("--enrich", metavar="PROJECT")
    ap.add_argument("--import-new", metavar="PROJECT", dest="import_new",
                    help="create entities for directory rows not already tracked, "
                         "tagged source=" + DIRECTORY_SOURCE)
    ap.add_argument("--apply", action="store_true",
                    help="with --enrich: actually write (default is a dry run)")
    ap.add_argument("--url", default=URL)
    a = ap.parse_args()
    if not a.out and not a.enrich and not a.import_new and not a.workbook:
        a.out = "out/nccde_associations.json"

    page = fetch(a.url)
    expected = re.search(r"of (\d+) Listing", page)
    recs = parse(page)

    # Do not let a silent parser regression look like a small county. If the
    # page says 224 and we parsed 12, that is a broken scrape, not a shrunken
    # directory -- exactly the "absence treated as valid" shape.
    if expected and len(recs) != int(expected.group(1)):
        print(f"REFUSING to write: page reports {expected.group(1)} listings, "
              f"parsed {len(recs)}. The page layout probably changed.",
              file=sys.stderr)
        return 2
    if not recs:
        print("REFUSING to write: parsed 0 listings.", file=sys.stderr)
        return 2

    if a.out:
        import os
        os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
        with open(a.out, "w") as f:
            json.dump(recs, f, indent=2)
        print(f"wrote {a.out}")

    managed = sum(r["managed"] for r in recs)
    print(f"{len(recs)} associations parsed")
    print(f"  self-managed (PROSPECTS): {len(recs) - managed}")
    print(f"  already professionally managed (competitors): {managed}")
    print(f"  with email: {sum(bool(r['email']) for r in recs)}"
          f" | with phone: {sum(bool(r['phone']) for r in recs)}")

    if a.workbook:
        print(f"workbook -> {workbook(recs, a.workbook)}")

    if a.import_new:
        print()
        import_new(a.import_new, recs, a.apply)
    if a.enrich:
        print()
        return enrich(a.enrich, recs, a.apply)
    return 0


if __name__ == "__main__":
    sys.exit(main())

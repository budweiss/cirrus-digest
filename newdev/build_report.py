#!/usr/bin/env python3
"""build_report.py — build DE-New-Developments.xlsx from out/plus_leads.json +
out/plus_owners.json (owner-enriched). Run AFTER plus_pull.py and plus_enrich.py.
Owner/Builder-of-record + mailing on every lead tab. (S47, 2026-07-29.)
Runs in the Cowork sandbox (openpyxl); reads/writes the property-management dir.
"""
import json, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
d = json.load(open(HERE / "out/plus_leads.json"))
da, bp, plus = d["dev_applications"], d["building_permits"], d["plus_projects"]
try:
    owners = json.load(open(HERE / "out/plus_owners.json"))
except FileNotFoundError:
    owners = {}


def own(a):
    o = owners.get((a.get("PARCEL_ID") or "").strip(), {})
    nm = o.get("owner", "")
    if o.get("owner2"):
        nm = (nm + " / " + o["owner2"]).strip(" /")
    return nm, o.get("mailing", "")


def clean(s):
    return re.sub(r"\s+", " ", (s or "").replace("\t", " ").replace("\n", " ")).strip()


def uni(a, k):
    u = a.get(k)
    return int(u) if isinstance(u, (int, float)) else (u or "")


def tier(u):
    u = u or 0
    return "A" if u >= 250 else "B" if u >= 150 else "C" if u >= 100 else "D"


SNOW = {"A": "$50k–150k+", "B": "$25k–60k", "C": "$15k–35k", "D": "$8k–20k"}
HDR = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HFILL = PatternFill("solid", fgColor="2F5496")
BODY = Font(name="Arial", size=10)
wrap = Alignment(vertical="top", wrap_text=True)
top = Alignment(vertical="top")
TF = {"A": PatternFill("solid", fgColor="C6E0B4"), "B": PatternFill("solid", fgColor="E2EFDA"),
      "C": PatternFill("solid", fgColor="FFF2CC"), "D": PatternFill("solid", fgColor="F2F2F2")}


def hdr(ws, H, W):
    for c, (h, w) in enumerate(zip(H, W), 1):
        cc = ws.cell(1, c, h); cc.font = HDR; cc.fill = HFILL; cc.alignment = wrap
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(H))}1"


wb = Workbook()
for a in da:
    a["_u"] = uni(a, "R_UNITS"); a["_t"] = tier(a["_u"])
das = sorted(da, key=lambda a: ((a.get("P_YEAR") or ""), a["_u"] if isinstance(a["_u"], int) else 0), reverse=True)
hot = [a for a in das if (a.get("P_YEAR") or "") in {"2022", "2023", "2024", "2025", "2026"}]


def devrow(ws, rows, extra=False):
    for a in rows:
        o, m = own(a)
        row = [a["_t"], a.get("P_YEAR", ""), a.get("COUNTY", ""), a["_u"], o, m,
               SNOW[a["_t"]], clean(a.get("NOTES")), a.get("RECTYPE", "")]
        if extra:
            row += [a.get("JURISDICTION", ""), a.get("PARCEL_ID", "")]
        r = ws.max_row + 1
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v); cell.font = BODY
            cell.alignment = wrap if c in (5, 8) else top
            if c == 1:
                cell.fill = TF[a["_t"]]; cell.font = Font(name="Arial", bold=True, size=10)


ws = wb.active; ws.title = "Top Targets"
H = ["Tier", "Year", "County", "Units", "Owner / Builder (of record)", "Owner mailing",
     "Est. snow (screen)", "Project / Location", "Record Type"]
hdr(ws, H, [5, 6, 13, 7, 30, 30, 17, 34, 16]); devrow(ws, hot)

ws2 = wb.create_sheet("All Dev Applications")
hdr(ws2, H + ["Jurisdiction", "Parcel ID"], [5, 6, 13, 7, 30, 30, 15, 32, 15, 18, 16])
devrow(ws2, das, extra=True)

ws3 = wb.create_sheet("Building Permits (active)")
H3 = ["Tier", "Year", "County", "Units", "Owner / Builder (of record)", "Owner mailing",
      "Est. snow (screen)", "Location / Notes", "Parcel ID"]
hdr(ws3, H3, [5, 6, 13, 7, 30, 30, 15, 32, 16])
for a in bp:
    u = uni(a, "R_UNITS"); t = tier(u); o, m = own(a)
    r = ws3.max_row + 1
    row = [t, a.get("P_YEAR", ""), a.get("COUNTY", ""), u, o, m, SNOW[t], clean(a.get("NOTES")), a.get("PARCEL_ID", "")]
    for c, v in enumerate(row, 1):
        cell = ws3.cell(r, c, v); cell.font = BODY; cell.alignment = wrap if c in (5, 8) else top
        if c == 1:
            cell.fill = TF[t]

ws4 = wb.create_sheet("PLUS Communities (built)")
H4 = ["Tier", "PLUS ID", "Year", "County", "Units", "Acres", "Owner / Applicant",
      "Engineer contact", "Est. snow", "Location", "Application"]
hdr(ws4, H4, [5, 12, 6, 13, 7, 6, 26, 20, 12, 34, 11])
for a in plus:
    u = uni(a, "RESIDENTIAL_UNITS"); t = tier(u); ac = uni(a, "PROJECT_AREA_ACRES")
    r = ws4.max_row + 1
    row = [t, a.get("PLUS_ID", ""), (a.get("PLUS_ID") or "")[:4], a.get("COUNTY", ""), u, ac,
           clean(a.get("OWNER_NAME")), clean(a.get("PROJECT_DESIGNER_NAME")), SNOW[t],
           clean(a.get("LOCATION")), "link" if a.get("APP_URL") else ""]
    for c, v in enumerate(row, 1):
        cell = ws4.cell(r, c, v); cell.font = BODY; cell.alignment = wrap if c in (7, 10) else top
        if c == 1:
            cell.fill = TF[t]
        if c == 11 and a.get("APP_URL"):
            cell.hyperlink = a["APP_URL"]; cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

ws0 = wb.create_sheet("How to use", 0); ws0.column_dimensions["A"].width = 116
matched = sum(1 for a in da + bp if own(a)[0])
L = [("DE New-Development Leads — Knight Property Services", 13, True),
     ("Source: DE OSPC planning data + Kent & Sussex county parcel/assessor systems. Refresh via runner plus-pull + plus-enrich.", 10, False),
     ("Filter: residential 50+ units, Kent + Sussex.", 10, False),
     ("", 10, False),
     ("OWNER / BUILDER (of record) = entity that owns the parcel (usually the developer LLC), from county assessor.", 10, True),
     (f"   Matched {matched} lead rows. Blank = parcel since subdivided; look up the Parcel ID at pride.kentcountyde.gov (Kent)", 10, False),
     ("   or map.sussexcountyde.gov (Sussex). OWNER MAILING = a direct contact address.", 10, False),
     ("TIERS by units: A=250+  B=150–249  C=100–149  D=50–99.  EST. SNOW = rough screen only, NOT a quote.", 10, False),
     ("Tabs: Top Targets (recent 2022–24) · All Dev Applications · Building Permits (building now) · PLUS Communities (built, at turnover).", 10, False),
     ("Approach + design: NEW-CONSTRUCTION-LEADS-DE.md.", 10, False)]
for i, (t, sz, b) in enumerate(L, 1):
    ws0.cell(i, 1, t).font = Font(name="Arial", size=sz, bold=b)

wb.save(HERE / "DE-New-Developments.xlsx")
print(f"built DE-New-Developments.xlsx | owner-matched lead rows: {matched} | tabs: {wb.sheetnames}")
